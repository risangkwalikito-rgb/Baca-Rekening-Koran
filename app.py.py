# file: app.py
from __future__ import annotations

import io
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="BCA Rekening Koran Reader", layout="wide")


COLUMN_ALIASES = {
    "trx_date": [
        "tanggal",
        "tgl",
        "date",
        "trx date",
        "transaction date",
        "posting date",
        "tanggal transaksi",
    ],
    "description": [
        "keterangan",
        "deskripsi",
        "uraian",
        "description",
        "transaction description",
        "remark",
        "remarks",
        "narrative",
    ],
    "debit": [
        "debit",
        "debet",
        "db",
    ],
    "credit": [
        "credit",
        "kredit",
        "cr",
    ],
    "balance": [
        "saldo",
        "balance",
        "running balance",
        "saldo akhir",
        "ending balance",
    ],
    "amount": [
        "mutasi",
        "amount",
        "nominal",
        "nilai",
    ],
    "dc": [
        "db/cr",
        "d/c",
        "dk",
        "type",
        "tipe",
        "jenis",
        "posisi",
        "mutasi type",
    ],
    "account_id": [
        "rekening",
        "no rekening",
        "no. rekening",
        "nomor rekening",
        "no rek",
        "norek",
        "rekening no",
        "account",
        "account no",
        "account number",
    ],
    "account_name": [
        "nama rekening",
        "nama account",
        "account name",
        "nama",
        "atas nama",
        "nama nasabah",
    ],
    "opening_balance_explicit": [
        "saldo awal",
        "opening balance",
        "beginning balance",
    ],
}


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def normalize_column_name(name: str) -> str:
    name = str(name or "").strip().lower()
    name = re.sub(r"[\r\n\t]+", " ", name)
    name = re.sub(r"[_\-]+", " ", name)
    name = re.sub(r"[^\w\s/\.]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def clean_account_id(value: object) -> str:
    raw = normalize_spaces(str(value or ""))
    digits = re.sub(r"\D", "", raw)
    if len(digits) >= 5:
        return digits
    return raw or "UNKNOWN"


def format_currency(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        num = float(value)
    except Exception:
        return str(value)
    return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def parse_amount(value: object) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = False
    if "(" in text and ")" in text:
        negative = True

    text = text.replace("Rp", "").replace("rp", "")
    text = text.replace("IDR", "").replace("idr", "")
    text = text.replace(" ", "")
    text = text.replace("\u00a0", "")
    text = re.sub(r"[^0-9,.\-]", "", text)

    if not text or text in {"-", ".", ","}:
        return None

    if text.startswith("-"):
        negative = True
        text = text[1:]

    if "." in text and "," in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts[-1]) == 2:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts[-1]) == 2 and len(parts) > 1:
            text = text.replace(",", "")
        else:
            text = text.replace(".", "")

    try:
        result = float(text)
    except ValueError:
        return None

    return -result if negative else result


def parse_date_value(value: object, year_hint: Optional[int] = None) -> pd.Timestamp:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return pd.NaT

    if isinstance(value, pd.Timestamp):
        return value.normalize()

    text = normalize_spaces(str(value))
    if not text:
        return pd.NaT

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            return pd.Timestamp(datetime.strptime(text, fmt).date())
        except ValueError:
            pass

    simple_dm = re.fullmatch(r"(\d{1,2})/(\d{1,2})", text)
    if simple_dm and year_hint:
        day = int(simple_dm.group(1))
        month = int(simple_dm.group(2))
        try:
            return pd.Timestamp(datetime(year_hint, month, day).date())
        except ValueError:
            return pd.NaT

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return pd.NaT
    return pd.Timestamp(parsed.date())


def guess_year_from_text(text: str) -> Optional[int]:
    years = re.findall(r"\b(20\d{2})\b", text)
    if years:
        counts = Counter(int(y) for y in years)
        return counts.most_common(1)[0][0]

    short_years = re.findall(r"\b\d{1,2}/\d{1,2}/(\d{2})\b", text)
    if short_years:
        inferred = [2000 + int(y) for y in short_years]
        counts = Counter(inferred)
        return counts.most_common(1)[0][0]

    return None


def detect_account_id_from_text(text: str, filename: str) -> str:
    patterns = [
        r"(?:NO\.?\s*REKENING|NOMOR\s*REKENING|NO\s*REK(?:ENING)?|ACCOUNT\s*NUMBER)\s*[:\-]?\s*([0-9 \-]{5,})",
        r"(?:REKENING)\s*[:\-]?\s*([0-9 \-]{5,})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return clean_account_id(match.group(1))

    return clean_account_id(Path(filename).stem)


def detect_account_name_from_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    patterns = [
        r"(?:NAMA\s*NASABAH|NAMA\s*REKENING|ATAS\s*NAMA|ACCOUNT\s*NAME)\s*[:\-]?\s*(.+)$",
        r"(?:NAMA)\s*[:\-]?\s*(.+)$",
    ]
    for line in lines:
        for pattern in patterns:
            match = re.search(pattern, line, flags=re.IGNORECASE)
            if match:
                candidate = normalize_spaces(match.group(1))
                candidate = re.sub(r"[^A-Za-z0-9 .,&/\-]", "", candidate).strip()
                if len(candidate) >= 3:
                    return candidate
    return ""


def detect_opening_balance_from_text(text: str) -> Optional[float]:
    patterns = [
        r"SALDO\s*AWAL\s*[:\-]?\s*([0-9.,]+)",
        r"OPENING\s*BALANCE\s*[:\-]?\s*([0-9.,]+)",
        r"BEGINNING\s*BALANCE\s*[:\-]?\s*([0-9.,]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return parse_amount(match.group(1))
    return None


def extract_pdf_text(file_bytes: bytes) -> str:
    texts: List[str] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            texts.append(page_text)
    return "\n".join(texts)


def merge_transaction_lines(lines: List[str]) -> List[str]:
    merged: List[str] = []
    current = ""

    for raw_line in lines:
        line = normalize_spaces(raw_line)
        if not line:
            continue

        if re.match(r"^\d{1,2}/\d{1,2}(?:/\d{2,4})?\b", line):
            if current:
                merged.append(current)
            current = line
        else:
            if current:
                current = f"{current} {line}"
            else:
                continue

    if current:
        merged.append(current)

    return merged


def detect_dc_marker(body: str, amount_span: Tuple[int, int]) -> Optional[str]:
    start, end = amount_span
    context = body[max(0, start - 20): min(len(body), end + 20)]

    if re.search(r"\b(?:DB|DEBIT|DEBET|D)\b", context, flags=re.IGNORECASE):
        return "DB"
    if re.search(r"\b(?:CR|KREDIT|CREDIT|K)\b", context, flags=re.IGNORECASE):
        return "CR"

    between_tail = body[end:]
    if re.search(r"^\s*(?:DB|DEBIT|DEBET|D)\b", between_tail, flags=re.IGNORECASE):
        return "DB"
    if re.search(r"^\s*(?:CR|KREDIT|CREDIT|K)\b", between_tail, flags=re.IGNORECASE):
        return "CR"

    before = body[:start]
    if re.search(r"\b(?:DB|DEBIT|DEBET|D)\s*$", before, flags=re.IGNORECASE):
        return "DB"
    if re.search(r"\b(?:CR|KREDIT|CREDIT|K)\s*$", before, flags=re.IGNORECASE):
        return "CR"

    return None


def parse_pdf_transaction_line(
    line: str,
    year_hint: Optional[int],
    account_id: str,
    account_name: str,
    opening_balance_explicit: Optional[float],
    source_file: str,
    row_order: int,
) -> Optional[Dict[str, object]]:
    line = normalize_spaces(line)
    date_match = re.match(r"^(\d{1,2}/\d{1,2}(?:/\d{2,4})?)\s+(.*)$", line)
    if not date_match:
        return None

    date_text = date_match.group(1)
    body = date_match.group(2).strip()

    if not body:
        return None

    money_pattern = re.compile(r"(?<!\d)(?:\d{1,3}(?:[.,]\d{3})+|\d+)(?:[.,]\d{2})?(?!\d)")
    matches = list(money_pattern.finditer(body))
    if len(matches) < 2:
        return None

    amount_match = matches[-2]
    balance_match = matches[-1]

    amount_text = amount_match.group(0)
    balance_text = balance_match.group(0)

    amount = parse_amount(amount_text)
    balance = parse_amount(balance_text)
    if amount is None and balance is None:
        return None

    dc_marker = detect_dc_marker(body, (amount_match.start(), amount_match.end()))

    description = body[: amount_match.start()].strip()
    description = re.sub(
        r"\b(?:DB|CR|DEBIT|DEBET|KREDIT|CREDIT|D|K)\s*$",
        "",
        description,
        flags=re.IGNORECASE,
    )
    description = normalize_spaces(description)

    if not description:
        description = "(tanpa keterangan)"

    debit = 0.0
    credit = 0.0
    if amount is not None:
        if dc_marker == "DB":
            debit = float(abs(amount))
        elif dc_marker == "CR":
            credit = float(abs(amount))

    return {
        "account_id": account_id,
        "account_name": account_name,
        "trx_date": parse_date_value(date_text, year_hint),
        "description": description,
        "amount": float(abs(amount)) if amount is not None else None,
        "debit": debit,
        "credit": credit,
        "balance": balance,
        "opening_balance_explicit": opening_balance_explicit,
        "source_file": source_file,
        "source_sheet": "PDF",
        "row_order": row_order,
        "dc_raw": dc_marker or "",
    }


def infer_missing_debit_credit(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.sort_values(["trx_date", "row_order"], kind="stable").copy()
    first_opening = df["opening_balance_explicit"].dropna()
    previous_balance = float(first_opening.iloc[0]) if not first_opening.empty else None

    for idx, row in df.iterrows():
        current_debit = float(row.get("debit", 0) or 0)
        current_credit = float(row.get("credit", 0) or 0)
        current_balance = row.get("balance")
        current_amount = row.get("amount")

        if current_debit > 0 or current_credit > 0:
            if pd.notna(current_balance):
                previous_balance = float(current_balance)
            continue

        if pd.notna(current_balance) and previous_balance is not None:
            delta = float(current_balance) - float(previous_balance)
            guessed_amount = float(abs(current_amount)) if pd.notna(current_amount) else abs(delta)

            if delta > 0:
                df.at[idx, "credit"] = guessed_amount
                df.at[idx, "debit"] = 0.0
            elif delta < 0:
                df.at[idx, "debit"] = guessed_amount
                df.at[idx, "credit"] = 0.0

            previous_balance = float(current_balance)
        elif pd.notna(current_amount):
            marker = str(row.get("dc_raw", "")).upper()
            if marker == "CR":
                df.at[idx, "credit"] = float(abs(current_amount))
                df.at[idx, "debit"] = 0.0
            elif marker == "DB":
                df.at[idx, "debit"] = float(abs(current_amount))
                df.at[idx, "credit"] = 0.0

            if pd.notna(current_balance):
                previous_balance = float(current_balance)
        elif pd.notna(current_balance):
            previous_balance = float(current_balance)

    return df


def read_csv_with_fallbacks(file_bytes: bytes) -> pd.DataFrame:
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1"]
    last_error: Optional[Exception] = None

    for encoding in encodings:
        try:
            return pd.read_csv(io.BytesIO(file_bytes), encoding=encoding)
        except Exception as exc:
            last_error = exc

    raise ValueError(f"Gagal membaca CSV: {last_error}")


def best_matching_column(columns: List[str], aliases: List[str]) -> Optional[str]:
    normalized_cols = {col: normalize_column_name(col) for col in columns}

    for alias in aliases:
        alias_norm = normalize_column_name(alias)
        for original, normalized in normalized_cols.items():
            if normalized == alias_norm:
                return original

    for alias in aliases:
        alias_norm = normalize_column_name(alias)
        for original, normalized in normalized_cols.items():
            if alias_norm in normalized or normalized in alias_norm:
                return original

    return None


def map_columns(df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    columns = list(df.columns)

    for canonical, aliases in COLUMN_ALIASES.items():
        matched = best_matching_column(columns, aliases)
        if matched:
            mapping[matched] = canonical

    return mapping


def standardize_dc(value: object) -> str:
    text = normalize_spaces(str(value or "")).upper()
    if text in {"DB", "DEBIT", "DEBET", "D"}:
        return "DB"
    if text in {"CR", "CREDIT", "KREDIT", "K"}:
        return "CR"
    return ""


def first_non_empty(series: pd.Series) -> str:
    for value in series:
        text = normalize_spaces(str(value or ""))
        if text:
            return text
    return ""


def create_manifest_row(
    account_id: str,
    account_name: str,
    source_file: str,
    source_sheet: str,
    parse_status: str,
    parse_note: str,
    transaction_count: int,
) -> Dict[str, object]:
    normalized_account_id = normalize_spaces(str(account_id or ""))
    if not normalized_account_id:
        fallback = Path(source_file).stem
        if source_sheet and source_sheet.upper() != "PDF":
            normalized_account_id = f"UNKNOWN::{fallback}::{source_sheet}"
        else:
            normalized_account_id = f"UNKNOWN::{fallback}"

    return {
        "account_id": normalized_account_id,
        "account_name": normalize_spaces(account_name),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "parse_status": parse_status,
        "parse_note": parse_note,
        "transaction_count": int(transaction_count),
    }


def build_manifest_from_transactions(
    parsed_df: pd.DataFrame,
    filename: str,
    sheet_name: str,
) -> pd.DataFrame:
    if parsed_df.empty:
        return pd.DataFrame(
            columns=[
                "account_id",
                "account_name",
                "source_file",
                "source_sheet",
                "parse_status",
                "parse_note",
                "transaction_count",
            ]
        )

    rows: List[Dict[str, object]] = []
    grouped = parsed_df.groupby("account_id", dropna=False, sort=True)

    for account_id, group in grouped:
        rows.append(
            create_manifest_row(
                account_id=str(account_id),
                account_name=first_non_empty(group["account_name"]) if "account_name" in group.columns else "",
                source_file=filename,
                source_sheet=sheet_name,
                parse_status="OK",
                parse_note="Transaksi berhasil dibaca",
                transaction_count=len(group),
            )
        )

    return pd.DataFrame(rows)


def extract_account_hints_from_dataframe(
    raw_df: pd.DataFrame,
    filename: str,
    sheet_name: str,
) -> List[Dict[str, str]]:
    df = raw_df.copy()
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

    if df.empty:
        return [
            {
                "account_id": f"UNKNOWN::{Path(filename).stem}::{sheet_name}",
                "account_name": "",
            }
        ]

    mapping = map_columns(df)
    df = df.rename(columns=mapping)

    if "account_id" in df.columns:
        temp = df.copy()
        temp["account_id"] = temp["account_id"].fillna("").astype(str).apply(normalize_spaces)
        if "account_name" in temp.columns:
            temp["account_name"] = temp["account_name"].fillna("").astype(str).apply(normalize_spaces)
        else:
            temp["account_name"] = ""

        unique_ids = []
        seen = set()

        for _, row in temp.iterrows():
            acc = normalize_spaces(row["account_id"])
            if not acc:
                continue

            acc_clean = clean_account_id(acc)
            if acc_clean in seen:
                continue

            seen.add(acc_clean)
            unique_ids.append(
                {
                    "account_id": acc_clean,
                    "account_name": normalize_spaces(row.get("account_name", "")),
                }
            )

        if unique_ids:
            return unique_ids

    flat_values = (
        df.astype(str)
        .replace("nan", "", regex=False)
        .replace("None", "", regex=False)
        .values
        .flatten()
        .tolist()
    )

    for value in flat_values:
        text = normalize_spaces(value)
        if not text:
            continue

        match = re.search(r"(?<!\d)(\d{8,20})(?!\d)", text)
        if match:
            return [
                {
                    "account_id": clean_account_id(match.group(1)),
                    "account_name": "",
                }
            ]

    return [
        {
            "account_id": f"UNKNOWN::{Path(filename).stem}::{sheet_name}",
            "account_name": "",
        }
    ]


def convert_spreadsheet_to_transactions(
    raw_df: pd.DataFrame,
    filename: str,
    sheet_name: str,
) -> Tuple[pd.DataFrame, List[str]]:
    notes: List[str] = []
    df = raw_df.copy()

    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.empty:
        return pd.DataFrame(), [f"{filename} [{sheet_name}]: sheet kosong"]

    mapping = map_columns(df)
    df = df.rename(columns=mapping)

    if "trx_date" not in df.columns:
        return pd.DataFrame(), [f"{filename} [{sheet_name}]: kolom tanggal tidak ditemukan"]

    if "description" not in df.columns:
        df["description"] = ""

    if "account_id" not in df.columns:
        fallback_account = sheet_name if sheet_name.lower() != "sheet1" else Path(filename).stem
        df["account_id"] = fallback_account

    if "account_name" not in df.columns:
        df["account_name"] = ""

    if "opening_balance_explicit" not in df.columns:
        df["opening_balance_explicit"] = None

    if "debit" not in df.columns:
        df["debit"] = None

    if "credit" not in df.columns:
        df["credit"] = None

    if "balance" not in df.columns:
        df["balance"] = None

    if "amount" not in df.columns:
        df["amount"] = None

    if "dc" not in df.columns:
        df["dc"] = ""

    df["trx_date"] = df["trx_date"].apply(parse_date_value)
    df["description"] = df["description"].fillna("").astype(str).apply(normalize_spaces)
    df["account_id"] = df["account_id"].apply(clean_account_id)
    df["account_name"] = df["account_name"].fillna("").astype(str).apply(normalize_spaces)
    df["opening_balance_explicit"] = df["opening_balance_explicit"].apply(parse_amount)
    df["debit"] = df["debit"].apply(parse_amount)
    df["credit"] = df["credit"].apply(parse_amount)
    df["balance"] = df["balance"].apply(parse_amount)
    df["amount"] = df["amount"].apply(parse_amount)
    df["dc"] = df["dc"].apply(standardize_dc)

    if df["amount"].notna().any():
        for idx, row in df.iterrows():
            amount = row["amount"]
            debit = row["debit"]
            credit = row["credit"]
            dc = row["dc"]

            if pd.isna(amount):
                continue

            if pd.isna(debit):
                debit = 0.0
            if pd.isna(credit):
                credit = 0.0

            if debit == 0 and credit == 0:
                if dc == "DB" or amount < 0:
                    df.at[idx, "debit"] = float(abs(amount))
                    df.at[idx, "credit"] = 0.0
                elif dc == "CR" or amount > 0:
                    df.at[idx, "credit"] = float(abs(amount))
                    df.at[idx, "debit"] = 0.0

    df["debit"] = df["debit"].fillna(0.0).astype(float)
    df["credit"] = df["credit"].fillna(0.0).astype(float)
    df["source_file"] = filename
    df["source_sheet"] = sheet_name
    df["row_order"] = range(len(df))
    df["dc_raw"] = df["dc"]

    required_cols = [
        "account_id",
        "account_name",
        "trx_date",
        "description",
        "amount",
        "debit",
        "credit",
        "balance",
        "opening_balance_explicit",
        "source_file",
        "source_sheet",
        "row_order",
        "dc_raw",
    ]
    for col in required_cols:
        if col not in df.columns:
            df[col] = None

    df = df[required_cols]
    df = df[df["trx_date"].notna()].copy()

    if df.empty:
        return pd.DataFrame(), [f"{filename} [{sheet_name}]: tidak ada baris transaksi valid"]

    notes.append(
        f"{filename} [{sheet_name}]: sheet terbaca | rekening unik={df['account_id'].nunique()} | transaksi={len(df)}"
    )
    return df, notes


def parse_bca_pdf(file_bytes: bytes, filename: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    notes: List[str] = []

    try:
        text = extract_pdf_text(file_bytes)
    except Exception as exc:
        manifest_df = pd.DataFrame(
            [
                create_manifest_row(
                    account_id=f"UNKNOWN::{Path(filename).stem}",
                    account_name="",
                    source_file=filename,
                    source_sheet="PDF",
                    parse_status="ERROR",
                    parse_note=f"Gagal baca PDF: {exc}",
                    transaction_count=0,
                )
            ]
        )
        return pd.DataFrame(), manifest_df, [f"{filename}: error baca PDF - {exc}"]

    if not text.strip():
        manifest_df = pd.DataFrame(
            [
                create_manifest_row(
                    account_id=f"UNKNOWN::{Path(filename).stem}",
                    account_name="",
                    source_file=filename,
                    source_sheet="PDF",
                    parse_status="NO_TEXT",
                    parse_note="PDF tidak mengandung teks yang bisa diekstrak",
                    transaction_count=0,
                )
            ]
        )
        return pd.DataFrame(), manifest_df, [f"{filename}: PDF tidak mengandung teks yang bisa diekstrak"]

    year_hint = guess_year_from_text(text)
    account_id = detect_account_id_from_text(text, filename)
    account_name = detect_account_name_from_text(text)
    opening_balance_explicit = detect_opening_balance_from_text(text)

    raw_lines = [line for line in text.splitlines() if line.strip()]
    merged_lines = merge_transaction_lines(raw_lines)

    rows: List[Dict[str, object]] = []
    for row_order, line in enumerate(merged_lines):
        parsed = parse_pdf_transaction_line(
            line=line,
            year_hint=year_hint,
            account_id=account_id,
            account_name=account_name,
            opening_balance_explicit=opening_balance_explicit,
            source_file=filename,
            row_order=row_order,
        )
        if parsed is not None:
            rows.append(parsed)

    if not rows:
        manifest_df = pd.DataFrame(
            [
                create_manifest_row(
                    account_id=account_id,
                    account_name=account_name,
                    source_file=filename,
                    source_sheet="PDF",
                    parse_status="NO_TRANSACTION",
                    parse_note="No rekening terdeteksi, tetapi transaksi tidak ditemukan",
                    transaction_count=0,
                )
            ]
        )
        return pd.DataFrame(), manifest_df, [f"{filename}: no rekening terdeteksi, transaksi PDF tidak ditemukan"]

    df = pd.DataFrame(rows)
    df = infer_missing_debit_credit(df)

    manifest_df = build_manifest_from_transactions(df, filename, "PDF")
    notes.append(f"{filename}: PDF terbaca | rekening={account_id} | transaksi={len(df)}")
    return df, manifest_df, notes


def parse_tabular_file(
    file_bytes: bytes,
    filename: str,
    ext: str,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    notes: List[str] = []
    results: List[pd.DataFrame] = []
    manifests: List[pd.DataFrame] = []

    if ext == ".csv":
        raw_df = read_csv_with_fallbacks(file_bytes)
        parsed_df, df_notes = convert_spreadsheet_to_transactions(raw_df, filename, "CSV")
        notes.extend(df_notes)

        if not parsed_df.empty:
            results.append(parsed_df)
            manifests.append(build_manifest_from_transactions(parsed_df, filename, "CSV"))
        else:
            hints = extract_account_hints_from_dataframe(raw_df, filename, "CSV")
            manifests.append(
                pd.DataFrame(
                    [
                        create_manifest_row(
                            account_id=hint["account_id"],
                            account_name=hint["account_name"],
                            source_file=filename,
                            source_sheet="CSV",
                            parse_status="NO_TRANSACTION",
                            parse_note="File terbaca tetapi tidak ada transaksi valid",
                            transaction_count=0,
                        )
                        for hint in hints
                    ]
                )
            )
    else:
        workbook = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)

        if not workbook:
            manifests.append(
                pd.DataFrame(
                    [
                        create_manifest_row(
                            account_id=f"UNKNOWN::{Path(filename).stem}",
                            account_name="",
                            source_file=filename,
                            source_sheet="WORKBOOK",
                            parse_status="EMPTY_WORKBOOK",
                            parse_note="Workbook kosong",
                            transaction_count=0,
                        )
                    ]
                )
            )

        for sheet_name, raw_df in workbook.items():
            parsed_df, df_notes = convert_spreadsheet_to_transactions(raw_df, filename, str(sheet_name))
            notes.extend(df_notes)

            if not parsed_df.empty:
                results.append(parsed_df)
                manifests.append(build_manifest_from_transactions(parsed_df, filename, str(sheet_name)))
            else:
                hints = extract_account_hints_from_dataframe(raw_df, filename, str(sheet_name))
                cleaned = raw_df.dropna(axis=0, how="all").dropna(axis=1, how="all")
                status = "EMPTY_SHEET" if cleaned.empty else "NO_TRANSACTION"
                note = "Sheet kosong" if cleaned.empty else "Sheet terbaca tetapi tidak ada transaksi valid"

                manifests.append(
                    pd.DataFrame(
                        [
                            create_manifest_row(
                                account_id=hint["account_id"],
                                account_name=hint["account_name"],
                                source_file=filename,
                                source_sheet=str(sheet_name),
                                parse_status=status,
                                parse_note=note,
                                transaction_count=0,
                            )
                            for hint in hints
                        ]
                    )
                )

    tx_df = pd.concat(results, ignore_index=True) if results else pd.DataFrame()
    manifest_df = pd.concat(manifests, ignore_index=True) if manifests else pd.DataFrame(
        columns=[
            "account_id",
            "account_name",
            "source_file",
            "source_sheet",
            "parse_status",
            "parse_note",
            "transaction_count",
        ]
    )
    return tx_df, manifest_df, notes


def finalize_transactions(df: pd.DataFrame, deduplicate: bool) -> pd.DataFrame:
    if df.empty:
        return df

    result = df.copy()
    result["account_id"] = result["account_id"].fillna("UNKNOWN").astype(str).apply(clean_account_id)
    result["account_name"] = result["account_name"].fillna("").astype(str).apply(normalize_spaces)
    result["description"] = result["description"].fillna("").astype(str).apply(normalize_spaces)
    result["debit"] = pd.to_numeric(result["debit"], errors="coerce").fillna(0.0)
    result["credit"] = pd.to_numeric(result["credit"], errors="coerce").fillna(0.0)
    result["balance"] = pd.to_numeric(result["balance"], errors="coerce")
    result["amount"] = pd.to_numeric(result["amount"], errors="coerce")
    result["opening_balance_explicit"] = pd.to_numeric(result["opening_balance_explicit"], errors="coerce")

    result = result.sort_values(
        by=["account_id", "trx_date", "source_file", "source_sheet", "row_order"],
        kind="stable",
    ).reset_index(drop=True)

    if deduplicate:
        result = result.drop_duplicates(
            subset=["account_id", "trx_date", "description", "debit", "credit", "balance"],
            keep="first",
        ).reset_index(drop=True)

    return result


def derive_opening_balance(group: pd.DataFrame) -> float:
    explicit = group["opening_balance_explicit"].dropna()
    if not explicit.empty:
        return float(explicit.iloc[0])

    first_valid_balance_rows = group[group["balance"].notna()]
    if not first_valid_balance_rows.empty:
        first = first_valid_balance_rows.iloc[0]
        return float(first["balance"]) + float(first["debit"]) - float(first["credit"])

    return 0.0


def derive_closing_balance(group: pd.DataFrame, opening_balance: float) -> float:
    valid_balances = group["balance"].dropna()
    if not valid_balances.empty:
        return float(valid_balances.iloc[-1])

    total_debit = float(group["debit"].sum())
    total_credit = float(group["credit"].sum())
    return opening_balance + total_credit - total_debit


def build_summary(df: pd.DataFrame, manifest_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    records: List[Dict[str, object]] = []

    if not df.empty:
        grouped = df.groupby("account_id", dropna=False, sort=True)

        for account_id, group in grouped:
            group = group.sort_values(
                by=["trx_date", "source_file", "source_sheet", "row_order"],
                kind="stable",
            ).reset_index(drop=True)

            opening_balance = derive_opening_balance(group)
            total_debit = float(group["debit"].sum())
            total_credit = float(group["credit"].sum())
            closing_balance = derive_closing_balance(group, opening_balance)
            account_name = first_non_empty(group["account_name"])

            records.append(
                {
                    "Rekening": str(account_id),
                    "Nama Rekening": account_name,
                    "Saldo Awal": opening_balance,
                    "Debit": total_debit,
                    "Kredit": total_credit,
                    "Saldo Akhir": closing_balance,
                    "Jumlah Transaksi": len(group),
                }
            )

    summary_df = pd.DataFrame(
        records,
        columns=[
            "Rekening",
            "Nama Rekening",
            "Saldo Awal",
            "Debit",
            "Kredit",
            "Saldo Akhir",
            "Jumlah Transaksi",
        ],
    )

    if manifest_df is not None and not manifest_df.empty:
        manifest_unique = (
            manifest_df[["account_id", "account_name"]]
            .fillna("")
            .drop_duplicates()
            .reset_index(drop=True)
        )

        existing_accounts = set(summary_df["Rekening"].astype(str)) if not summary_df.empty else set()
        missing_rows: List[Dict[str, object]] = []

        for _, row in manifest_unique.iterrows():
            account_id = normalize_spaces(str(row["account_id"]))
            account_name = normalize_spaces(str(row["account_name"]))
            if account_id in existing_accounts:
                continue

            missing_rows.append(
                {
                    "Rekening": account_id,
                    "Nama Rekening": account_name,
                    "Saldo Awal": 0.0,
                    "Debit": 0.0,
                    "Kredit": 0.0,
                    "Saldo Akhir": 0.0,
                    "Jumlah Transaksi": 0,
                }
            )

        if missing_rows:
            summary_df = pd.concat([summary_df, pd.DataFrame(missing_rows)], ignore_index=True)

    if summary_df.empty:
        summary_df = pd.DataFrame(
            columns=[
                "Rekening",
                "Nama Rekening",
                "Saldo Awal",
                "Debit",
                "Kredit",
                "Saldo Akhir",
                "Jumlah Transaksi",
            ]
        )

    summary_df = summary_df.sort_values(by="Rekening", kind="stable").reset_index(drop=True)
    return summary_df


def make_display_copy(df: pd.DataFrame, money_columns: List[str]) -> pd.DataFrame:
    display_df = df.copy()
    for col in money_columns:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_currency)
    return display_df


def sanitize_sheet_name(name: str, used_names: set) -> str:
    clean = re.sub(r"[:\\/?*\[\]]", "_", str(name)).strip()
    clean = clean[:31] or "Sheet"

    if clean not in used_names:
        used_names.add(clean)
        return clean

    base = clean[:28] or "Sht"
    counter = 1
    while True:
        candidate = f"{base}_{counter}"[:31]
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def build_excel_split_by_date(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    manifest_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()
    used_sheet_names = set()

    detail_export = detail_df.copy()
    if "Tanggal" in detail_export.columns:
        detail_export["Tanggal"] = pd.to_datetime(detail_export["Tanggal"], errors="coerce")

    status_export = manifest_df.copy()
    if status_export.empty:
        status_export = pd.DataFrame(
            columns=[
                "account_id",
                "account_name",
                "source_file",
                "source_sheet",
                "parse_status",
                "parse_note",
                "transaction_count",
            ]
        )

    status_export = status_export.rename(
        columns={
            "account_id": "Rekening",
            "account_name": "Nama Rekening",
            "source_file": "File",
            "source_sheet": "Sheet",
            "parse_status": "Status",
            "parse_note": "Catatan",
            "transaction_count": "Jumlah Transaksi",
        }
    )

    status_export = status_export.sort_values(
        by=["File", "Sheet", "Rekening"],
        kind="stable",
    ).reset_index(drop=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_export = summary_df.copy()
        summary_export.to_excel(
            writer,
            sheet_name=sanitize_sheet_name("Rekap", used_sheet_names),
            index=False,
        )

        all_export = detail_export.copy()
        if "Tanggal" in all_export.columns:
            all_export["Tanggal"] = all_export["Tanggal"].dt.strftime("%Y-%m-%d")

        all_export.to_excel(
            writer,
            sheet_name=sanitize_sheet_name("Semua_Transaksi", used_sheet_names),
            index=False,
        )

        status_export.to_excel(
            writer,
            sheet_name=sanitize_sheet_name("Status_File", used_sheet_names),
            index=False,
        )

        if not detail_export.empty and "Tanggal" in detail_export.columns:
            dated_rows = detail_export[detail_export["Tanggal"].notna()].copy()
            if not dated_rows.empty:
                dated_rows = dated_rows.sort_values(
                    ["Tanggal", "Rekening", "File", "Sheet"],
                    kind="stable",
                )

                for trx_date, group in dated_rows.groupby(dated_rows["Tanggal"].dt.date, sort=True):
                    sheet_name = sanitize_sheet_name(str(trx_date), used_sheet_names)
                    export_group = group.copy()
                    export_group["Tanggal"] = export_group["Tanggal"].dt.strftime("%Y-%m-%d")
                    export_group.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        money_columns = {"Saldo Awal", "Debit", "Kredit", "Saldo Akhir", "Saldo"}

        for ws in workbook.worksheets:
            headers = [cell.value for cell in ws[1]]

            for col_idx, header in enumerate(headers, start=1):
                if header in money_columns:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

            autosize_worksheet(ws)

    output.seek(0)
    return output.getvalue()


def main() -> None:
    st.title("BCA Rekening Koran Reader")
    st.caption("Upload banyak file sekaligus, gabungkan banyak rekening, lalu rekap per rekening.")

    with st.sidebar:
        st.subheader("Opsi")
        deduplicate = st.checkbox("Hapus duplikat transaksi identik", value=True)
        st.markdown(
            """
            **Format file yang didukung**
            - PDF rekening koran / mutasi BCA berbasis teks
            - CSV
            - XLSX / XLS

            **Kolom spreadsheet yang dikenali**
            - tanggal
            - keterangan
            - debit
            - kredit
            - saldo
            - rekening
            - nama rekening
            - saldo awal
            """
        )

    uploaded_files = st.file_uploader(
        "Pilih file rekening koran / mutasi",
        type=["pdf", "csv", "xlsx", "xls"],
        accept_multiple_files=True,
    )

    if not uploaded_files:
        st.info("Upload file dulu untuk mulai proses.")
        return

    parsed_dfs: List[pd.DataFrame] = []
    manifest_dfs: List[pd.DataFrame] = []
    parser_notes: List[str] = []
    parser_errors: List[str] = []

    progress = st.progress(0, text="Memproses file...")
    total_files = len(uploaded_files)

    for i, uploaded_file in enumerate(uploaded_files, start=1):
        filename = uploaded_file.name
        ext = Path(filename).suffix.lower()
        file_bytes = uploaded_file.getvalue()

        try:
            if ext == ".pdf":
                df_file, manifest_file, notes = parse_bca_pdf(file_bytes, filename)
            elif ext in {".csv", ".xlsx", ".xls"}:
                df_file, manifest_file, notes = parse_tabular_file(file_bytes, filename, ext)
            else:
                df_file = pd.DataFrame()
                manifest_file = pd.DataFrame(
                    [
                        create_manifest_row(
                            account_id=f"UNKNOWN::{Path(filename).stem}",
                            account_name="",
                            source_file=filename,
                            source_sheet="FILE",
                            parse_status="UNSUPPORTED",
                            parse_note="Format file tidak didukung",
                            transaction_count=0,
                        )
                    ]
                )
                notes = [f"{filename}: format file tidak didukung"]

            parser_notes.extend(notes)

            if not manifest_file.empty:
                manifest_dfs.append(manifest_file)

            if not df_file.empty:
                parsed_dfs.append(df_file)
            else:
                parser_errors.append(f"{filename}: tidak ada transaksi valid, tetapi file tetap dicatat")

        except Exception as exc:
            parser_errors.append(f"{filename}: error - {exc}")
            manifest_dfs.append(
                pd.DataFrame(
                    [
                        create_manifest_row(
                            account_id=f"UNKNOWN::{Path(filename).stem}",
                            account_name="",
                            source_file=filename,
                            source_sheet="FILE",
                            parse_status="ERROR",
                            parse_note=str(exc),
                            transaction_count=0,
                        )
                    ]
                )
            )

        progress.progress(i / total_files, text=f"Memproses file {i}/{total_files}: {filename}")

    progress.empty()

    if not manifest_dfs:
        st.error("Tidak ada file yang berhasil dicatat.")
        return

    manifest_df = pd.concat(manifest_dfs, ignore_index=True)
    transactions = pd.concat(parsed_dfs, ignore_index=True) if parsed_dfs else pd.DataFrame()

    if not transactions.empty:
        transactions = finalize_transactions(transactions, deduplicate=deduplicate)

    summary = build_summary(transactions, manifest_df=manifest_df)

    st.success(
        f"Upload: {len(uploaded_files)} file | "
        f"Tercatat di hasil: {manifest_df['source_file'].nunique()} file | "
        f"Dengan transaksi: {len(parsed_dfs)} file"
    )

    st.subheader("Rekap per Rekening")
    summary_display = make_display_copy(
        summary,
        money_columns=["Saldo Awal", "Debit", "Kredit", "Saldo Akhir"],
    )
    st.dataframe(summary_display, use_container_width=True, hide_index=True)

    st.subheader("Status File / Sheet")
    status_preview = manifest_df.rename(
        columns={
            "account_id": "Rekening",
            "account_name": "Nama Rekening",
            "source_file": "File",
            "source_sheet": "Sheet",
            "parse_status": "Status",
            "parse_note": "Catatan",
            "transaction_count": "Jumlah Transaksi",
        }
    )
    st.dataframe(status_preview, use_container_width=True, hide_index=True)

    if not transactions.empty:
        detail_columns = [
            "account_id",
            "account_name",
            "trx_date",
            "description",
            "debit",
            "credit",
            "balance",
            "source_file",
            "source_sheet",
        ]
        detail_df = transactions[detail_columns].rename(
            columns={
                "account_id": "Rekening",
                "account_name": "Nama Rekening",
                "trx_date": "Tanggal",
                "description": "Keterangan",
                "debit": "Debit",
                "credit": "Kredit",
                "balance": "Saldo",
                "source_file": "File",
                "source_sheet": "Sheet",
            }
        )
    else:
        detail_df = pd.DataFrame(
            columns=[
                "Rekening",
                "Nama Rekening",
                "Tanggal",
                "Keterangan",
                "Debit",
                "Kredit",
                "Saldo",
                "File",
                "Sheet",
            ]
        )

    detail_display = make_display_copy(detail_df, money_columns=["Debit", "Kredit", "Saldo"])
    st.subheader("Detail Transaksi")
    st.dataframe(detail_display, use_container_width=True, hide_index=True)

    excel_bytes = build_excel_split_by_date(summary, detail_df, manifest_df)

    st.download_button(
        label="Download Excel Split per Tanggal",
        data=excel_bytes,
        file_name="rekap_bca_split_per_tanggal.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        label="Download Rekap CSV",
        data=summary.to_csv(index=False).encode("utf-8-sig"),
        file_name="rekap_rekening_bca.csv",
        mime="text/csv",
    )

    st.download_button(
        label="Download Detail CSV",
        data=detail_df.to_csv(index=False).encode("utf-8-sig"),
        file_name="detail_transaksi_bca.csv",
        mime="text/csv",
    )

    st.download_button(
        label="Download Status File CSV",
        data=status_preview.to_csv(index=False).encode("utf-8-sig"),
        file_name="status_file_bca.csv",
        mime="text/csv",
    )

    with st.expander("Detail log parser"):
        for note in parser_notes:
            st.write(f"- {note}")
        for err in parser_errors:
            st.write(f"- {err}")


if __name__ == "__main__":
    main()
