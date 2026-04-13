# file: app.py
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Statement Reader", layout="wide")


MASTER_BCA = [
    ("0613419702", "BANK BCA CASHLESS BATAM"),
    ("4301191191", "BANK BCA PNP BAKAUHENI"),
    ("8200827831", "BANK BCA CASHLESS SIBOLGA"),
    ("2950400652", "BANK BCA PNP MERAK"),
    ("2642537777", "BANK BCA PNP KETAPANG"),
    ("1870888828", "BANK BCA PNP SURABAYA"),
    ("7810337154", "BANK BCA CASHLESS BALIKPAPAN"),
    ("8685126334", "BANK BCA CASHLESS BATULICIN"),
    ("7855301644", "BANK BCA CASHLESS TERNATE"),
    ("3141086306", "BANK BCA CASHLESS KUPANG"),
    ("7255999001", "BANK BCA PNP KAYANGAN"),
    ("0561743893", "BANK BCA PNP LEMBAR"),
    ("7065038676", "BANK BCA CASHLESS SAPE"),
    ("8745194440", "BANK BCA CASHLESS BAJOE"),
    ("0411613436", "BANK BCA BANGKA"),
    ("3900925572", "BANK BCA SELAYAR"),
    ("0441598776", "BANK BCA AMBON"),
    ("0223259861", "BANK BCA ACEH"),
    ("0322725645", "BANK BCA PADANG"),
    ("6795136821", "BANK BCA LUWUK"),
    ("6495342828", "BANK BCA BAU-BAU"),
    ("4500553842", "BANK BCA PELABUHAN"),
]

MASTER_BY_BANK = {
    "BCA": MASTER_BCA,
    "MANDIRI": [],
    "BNI": [],
    "LAINNYA": [],
}

ACCOUNT_LENGTH_BY_BANK = {
    "BCA": 10,
    "MANDIRI": 13,
    "BNI": 10,
    "LAINNYA": 10,
}

HEADER_ALIASES = {
    "trx_date": [
        "tanggal", "tgl", "date", "posting date", "transaction date",
    ],
    "description": [
        "keterangan", "deskripsi", "uraian", "remark", "remarks", "description", "narrative",
    ],
    "account_id": [
        "rekening", "no rekening", "no. rekening", "nomor rekening", "no rek", "norek",
        "account", "account no", "account number", "account no.",
    ],
    "account_name": [
        "nama rekening", "nama account", "account name", "nama", "atas nama", "nama nasabah",
    ],
    "opening_balance_explicit": [
        "saldo awal", "opening balance", "beginning balance",
    ],
    "debit": ["debit", "debet", "db"],
    "credit": ["credit", "kredit", "cr"],
    "balance": ["saldo", "balance", "running balance", "ending balance", "saldo akhir"],
    "amount": ["mutasi", "amount", "nominal", "nilai"],
    "dc": ["jenis transaksi", "db/cr", "d/c", "dk", "type", "tipe", "jenis", "posisi", "mutasi type"],
}

STATUS_COLUMNS = [
    "account_key", "account_name", "source_file", "source_sheet", "parse_status", "parse_note", "transaction_count",
]


def normalize_spaces(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_col(value: object) -> str:
    text = normalize_spaces(value).lower()
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"[^\w\s/\.]", " ", text)
    return normalize_spaces(text)


def get_account_length(bank: str) -> int:
    return ACCOUNT_LENGTH_BY_BANK.get(bank.upper(), 10)


def normalize_account_key(value: object, bank: str) -> str:
    length = get_account_length(bank)
    digits = re.sub(r"\D", "", str(value or ""))
    if digits:
        return digits[:length]
    text = normalize_spaces(value)
    return text[:length] if text else ""


def format_currency(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        number = float(value)
    except Exception:
        return str(value)
    return f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def parse_amount(value: object) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)

    text = normalize_spaces(value)
    if not text:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    text = (
        text.replace("Rp", "")
        .replace("rp", "")
        .replace("IDR", "")
        .replace("idr", "")
        .replace(" ", "")
        .replace("\u00a0", "")
    )
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text:
        return None

    if text.startswith("-"):
        negative = True
        text = text[1:]

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts[-1]) == 2:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts[-1]) != 2:
            text = text.replace(".", "")

    try:
        result = float(text)
    except Exception:
        return None
    return -result if negative else result


def parse_date_value(value: object) -> pd.Timestamp:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value

    text = normalize_spaces(value)
    if not text:
        return pd.NaT

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return pd.NaT
    return pd.Timestamp(parsed)


def standardize_dc(value: object) -> str:
    text = normalize_spaces(value).upper()
    if text in {"DB", "DEBIT", "DEBET", "D"}:
        return "DB"
    if text in {"CR", "CREDIT", "KREDIT", "K"}:
        return "CR"
    return ""


def create_status_row(
    bank: str,
    account_key: str,
    account_name: str,
    source_file: str,
    source_sheet: str,
    parse_status: str,
    parse_note: str,
    transaction_count: int,
) -> Dict[str, object]:
    key = normalize_account_key(account_key, bank)
    if not key:
        key = f"UNKNOWN::{Path(source_file).stem}"
        if source_sheet:
            key = f"{key}::{source_sheet}"

    return {
        "account_key": key,
        "account_name": normalize_spaces(account_name),
        "source_file": source_file,
        "source_sheet": source_sheet,
        "parse_status": parse_status,
        "parse_note": parse_note,
        "transaction_count": int(transaction_count),
    }


def build_master_dataframe(bank: str) -> pd.DataFrame:
    rows = []
    for order_no, (account_key, account_name) in enumerate(MASTER_BY_BANK.get(bank.upper(), []), start=1):
        rows.append(
            {
                "account_key": normalize_account_key(account_key, bank),
                "account_name": account_name,
                "master_order": order_no,
            }
        )
    return pd.DataFrame(rows)


def best_matching_column(columns: List[str], aliases: List[str]) -> Optional[str]:
    normalized = {col: normalize_col(col) for col in columns}
    for alias in aliases:
        alias_n = normalize_col(alias)
        for original, current in normalized.items():
            if current == alias_n:
                return original
    for alias in aliases:
        alias_n = normalize_col(alias)
        for original, current in normalized.items():
            if alias_n in current or current in alias_n:
                return original
    return None


def detect_header_row(raw_df: pd.DataFrame) -> int:
    best_row = 0
    best_score = -1
    target_words = [alias for values in HEADER_ALIASES.values() for alias in values]

    scan_limit = min(len(raw_df), 25)
    for idx in range(scan_limit):
        row_texts = [normalize_col(v) for v in raw_df.iloc[idx].tolist()]
        score = 0
        for cell in row_texts:
            if not cell:
                continue
            for word in target_words:
                word_n = normalize_col(word)
                if cell == word_n or word_n in cell:
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = idx
    return best_row


def extract_account_hints_from_raw(raw_df: pd.DataFrame, bank: str, filename: str, sheet_name: str) -> List[Tuple[str, str]]:
    flat_values = [normalize_spaces(v) for v in raw_df.astype(str).fillna("").values.flatten().tolist()]
    length = get_account_length(bank)
    candidates: List[Tuple[str, str]] = []

    for value in flat_values[:200]:
        digits = re.sub(r"\D", "", value)
        if len(digits) >= length:
            key = normalize_account_key(digits, bank)
            if key:
                candidates.append((key, ""))

    unique: List[Tuple[str, str]] = []
    seen = set()
    for key, name in candidates:
        if key in seen:
            continue
        seen.add(key)
        unique.append((key, name))
    if unique:
        return unique

    fallback = normalize_account_key(Path(filename).stem, bank) or f"UNKNOWN::{Path(filename).stem}::{sheet_name}"
    return [(fallback, "")]


def extract_statement_text(file_bytes: bytes) -> str:
    parts: List[str] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def detect_bca_account_id(text: str, filename: str, bank: str) -> str:
    patterns = [
        r"(?:NO\.?\s*REKENING|NOMOR\s*REKENING|NO\s*REK(?:ENING)?)\s*[:\-]?\s*([0-9 \-]{5,})",
        r"(?:REKENING)\s*[:\-]?\s*([0-9 \-]{5,})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return normalize_account_key(match.group(1), bank)
    return normalize_account_key(Path(filename).stem, bank)


def detect_bca_account_name(text: str) -> str:
    for line in text.splitlines():
        line_clean = normalize_spaces(line)
        match = re.search(r"(?:NAMA\s*REKENING|NAMA\s*NASABAH|ATAS\s*NAMA)\s*[:\-]?\s*(.+)$", line_clean, flags=re.IGNORECASE)
        if match:
            return normalize_spaces(match.group(1))
    return ""


def detect_opening_balance(text: str) -> Optional[float]:
    for pattern in [
        r"SALDO\s*AWAL\s*[:\-]?\s*([0-9.,]+)",
        r"OPENING\s*BALANCE\s*[:\-]?\s*([0-9.,]+)",
        r"BEGINNING\s*BALANCE\s*[:\-]?\s*([0-9.,]+)",
    ]:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return parse_amount(match.group(1))
    return None


def merge_bca_lines(lines: List[str]) -> List[str]:
    merged: List[str] = []
    current = ""
    for raw_line in lines:
        line = normalize_spaces(raw_line)
        if not line:
            continue
        if re.match(r"^\d{1,2}/\d{1,2}(?:/\d{2,4})?\b", line):
            if current:
                merged.append(current)
            current = line
        else:
            if current:
                current = f"{current} {line}"
    if current:
        merged.append(current)
    return merged


def parse_bca_pdf(file_bytes: bytes, filename: str, bank: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    notes: List[str] = []
    try:
        text = extract_statement_text(file_bytes)
    except Exception as exc:
        status = pd.DataFrame([create_status_row(bank, "", "", filename, "PDF", "ERROR", f"Gagal baca PDF: {exc}", 0)])
        return pd.DataFrame(), status, [f"{filename}: {exc}"]

    if not text.strip():
        status = pd.DataFrame([create_status_row(bank, "", "", filename, "PDF", "NO_TEXT", "PDF tidak punya text layer", 0)])
        return pd.DataFrame(), status, [f"{filename}: PDF tidak punya text layer"]

    account_key = detect_bca_account_id(text, filename, bank)
    account_name = detect_bca_account_name(text)
    opening_balance = detect_opening_balance(text)
    lines = merge_bca_lines([x for x in text.splitlines() if normalize_spaces(x)])

    rows: List[Dict[str, object]] = []
    money_pattern = re.compile(r"(?<!\d)(?:\d{1,3}(?:[.,]\d{3})+|\d+)(?:[.,]\d{2})?(?!\d)")

    for row_order, line in enumerate(lines):
        match = re.match(r"^(\d{1,2}/\d{1,2}(?:/\d{2,4})?)\s+(.*)$", line)
        if not match:
            continue
        date_text, body = match.group(1), match.group(2)
        amounts = list(money_pattern.finditer(body))
        if len(amounts) < 2:
            continue

        amount_text = amounts[-2].group(0)
        balance_text = amounts[-1].group(0)
        amount = parse_amount(amount_text)
        balance = parse_amount(balance_text)
        desc = normalize_spaces(body[:amounts[-2].start()])

        context = body[max(0, amounts[-2].start() - 20): min(len(body), amounts[-2].end() + 20)]
        dc = ""
        if re.search(r"\b(?:DB|DEBIT|DEBET|D)\b", context, flags=re.IGNORECASE):
            dc = "DB"
        elif re.search(r"\b(?:CR|KREDIT|CREDIT|K)\b", context, flags=re.IGNORECASE):
            dc = "CR"

        debit = 0.0
        credit = 0.0
        if amount is not None:
            if dc == "DB":
                debit = abs(float(amount))
            elif dc == "CR":
                credit = abs(float(amount))

        rows.append(
            {
                "bank": bank,
                "account_key": account_key,
                "account_name": account_name,
                "trx_date": parse_date_value(date_text).normalize(),
                "description": desc,
                "amount": abs(float(amount)) if amount is not None else None,
                "dc": dc,
                "debit": debit,
                "credit": credit,
                "balance": balance,
                "opening_balance_explicit": opening_balance,
                "source_file": filename,
                "source_sheet": "PDF",
                "row_order": row_order,
            }
        )

    if not rows:
        status = pd.DataFrame([create_status_row(bank, account_key, account_name, filename, "PDF", "NO_TRANSACTION", "No rekening terbaca, transaksi tidak ditemukan", 0)])
        return pd.DataFrame(), status, [f"{filename}: transaksi tidak ditemukan"]

    df = pd.DataFrame(rows)
    status = pd.DataFrame([create_status_row(bank, account_key, account_name, filename, "PDF", "OK", "Transaksi berhasil dibaca", len(df))])
    notes.append(f"{filename}: PDF terbaca | rekening={account_key} | transaksi={len(df)}")
    return df, status, notes


def parse_mandiri_sheet(raw_df: pd.DataFrame, filename: str, sheet_name: str, bank: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    notes: List[str] = []
    if raw_df.empty:
        status = pd.DataFrame([create_status_row(bank, "", "", filename, sheet_name, "EMPTY_SHEET", "Sheet kosong", 0)])
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: kosong"]

    account_cell = normalize_spaces(raw_df.iloc[5, 6]) if raw_df.shape[0] > 5 and raw_df.shape[1] > 6 else ""
    opening_cell = raw_df.iloc[9, 6] if raw_df.shape[0] > 9 and raw_df.shape[1] > 6 else None

    account_key = normalize_account_key(account_cell, bank)
    account_name = ""
    if account_cell:
        m = re.search(rf"({re.escape(account_key)})\s+(?:IDR|USD|SGD)?\s*(.*)$", account_cell)
        if m:
            account_name = normalize_spaces(m.group(2))

    opening_balance = parse_amount(opening_cell)

    if raw_df.shape[0] <= 11:
        status = pd.DataFrame([create_status_row(bank, account_key, account_name, filename, sheet_name, "NO_TRANSACTION", "Format Mandiri tidak lengkap", 0)])
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: format tidak lengkap"]

    data = raw_df.iloc[12:, :].copy()
    if data.empty:
        status = pd.DataFrame([create_status_row(bank, account_key, account_name, filename, sheet_name, "NO_TRANSACTION", "Tidak ada data transaksi", 0)])
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: tidak ada data"]

    rows: List[Dict[str, object]] = []
    for i, (_, row) in enumerate(data.iterrows()):
        posting_date = parse_date_value(row.iloc[1] if len(row) > 1 else None)
        if pd.isna(posting_date):
            continue

        description = normalize_spaces(row.iloc[4] if len(row) > 4 else "")
        debit = parse_amount(row.iloc[9] if len(row) > 9 else None) or 0.0
        credit = parse_amount(row.iloc[11] if len(row) > 11 else None) or 0.0
        balance = parse_amount(row.iloc[15] if len(row) > 15 else None)

        if balance is None and debit == 0 and credit == 0 and not description:
            continue

        dc = ""
        amount = None
        if credit > 0:
            dc = "CR"
            amount = credit
        elif debit > 0:
            dc = "DB"
            amount = debit

        rows.append(
            {
                "bank": bank,
                "account_key": account_key,
                "account_name": account_name,
                "trx_date": posting_date.normalize(),
                "description": description,
                "amount": amount,
                "dc": dc,
                "debit": float(debit),
                "credit": float(credit),
                "balance": balance,
                "opening_balance_explicit": opening_balance,
                "source_file": filename,
                "source_sheet": sheet_name,
                "row_order": i,
            }
        )

    if not rows:
        status = pd.DataFrame([create_status_row(bank, account_key, account_name, filename, sheet_name, "NO_TRANSACTION", "Tidak ada transaksi valid", 0)])
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: tidak ada transaksi valid"]

    df = pd.DataFrame(rows)
    status = pd.DataFrame([create_status_row(bank, account_key, account_name, filename, sheet_name, "OK", "Transaksi Mandiri berhasil dibaca", len(df))])
    notes.append(f"{filename}[{sheet_name}]: rekening={account_key} | transaksi={len(df)}")
    return df, status, notes


def prepare_standardized_frame(raw_df: pd.DataFrame) -> pd.DataFrame:
    raw_df = raw_df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if raw_df.empty:
        return raw_df

    header_row = detect_header_row(raw_df)
    header = [normalize_spaces(v) or f"col_{i}" for i, v in enumerate(raw_df.iloc[header_row].tolist())]
    data = raw_df.iloc[header_row + 1 :].copy()
    data.columns = header
    data = data.dropna(axis=0, how="all")
    return data


def map_standard_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping: Dict[str, str] = {}
    columns = list(df.columns)
    for target, aliases in HEADER_ALIASES.items():
        found = best_matching_column(columns, aliases)
        if found:
            mapping[found] = target
    return df.rename(columns=mapping)


def parse_generic_tabular_sheet(
    raw_df: pd.DataFrame,
    filename: str,
    sheet_name: str,
    bank: str,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    notes: List[str] = []

    standardized = prepare_standardized_frame(raw_df)
    if standardized.empty:
        hints = extract_account_hints_from_raw(raw_df, bank, filename, sheet_name)
        status = pd.DataFrame(
            [create_status_row(bank, key, name, filename, sheet_name, "EMPTY_SHEET", "Sheet kosong", 0) for key, name in hints]
        )
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: kosong"]

    df = map_standard_columns(standardized)

    if "trx_date" not in df.columns:
        hints = extract_account_hints_from_raw(raw_df, bank, filename, sheet_name)
        status = pd.DataFrame(
            [create_status_row(bank, key, name, filename, sheet_name, "NO_DATE_COLUMN", "Kolom tanggal tidak ditemukan", 0) for key, name in hints]
        )
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: kolom tanggal tidak ditemukan"]

    if "description" not in df.columns:
        df["description"] = ""

    if "account_id" in df.columns:
        df["account_key"] = df["account_id"].apply(lambda x: normalize_account_key(x, bank))
    else:
        hints = extract_account_hints_from_raw(raw_df, bank, filename, sheet_name)
        fallback_key = hints[0][0]
        df["account_key"] = fallback_key

    if "account_name" not in df.columns:
        df["account_name"] = ""

    if "debit" not in df.columns:
        df["debit"] = None
    if "credit" not in df.columns:
        df["credit"] = None
    if "balance" not in df.columns:
        df["balance"] = None
    if "amount" not in df.columns:
        df["amount"] = None
    if "dc" not in df.columns:
        df["dc"] = ""
    if "opening_balance_explicit" not in df.columns:
        df["opening_balance_explicit"] = None

    df["trx_date"] = df["trx_date"].apply(parse_date_value).dt.normalize()
    df["description"] = df["description"].apply(normalize_spaces)
    df["account_name"] = df["account_name"].apply(normalize_spaces)
    df["debit"] = df["debit"].apply(parse_amount)
    df["credit"] = df["credit"].apply(parse_amount)
    df["balance"] = df["balance"].apply(parse_amount)
    df["amount"] = df["amount"].apply(parse_amount)
    df["dc"] = df["dc"].apply(standardize_dc)
    df["opening_balance_explicit"] = df["opening_balance_explicit"].apply(parse_amount)

    for idx, row in df.iterrows():
        debit = row["debit"] if pd.notna(row["debit"]) else 0.0
        credit = row["credit"] if pd.notna(row["credit"]) else 0.0
        amount = row["amount"]
        dc = row["dc"]

        if (debit == 0 or pd.isna(debit)) and (credit == 0 or pd.isna(credit)) and pd.notna(amount):
            if dc == "CR" or amount > 0:
                df.at[idx, "credit"] = abs(float(amount))
                df.at[idx, "debit"] = 0.0
            elif dc == "DB" or amount < 0:
                df.at[idx, "debit"] = abs(float(amount))
                df.at[idx, "credit"] = 0.0

    df["debit"] = pd.to_numeric(df["debit"], errors="coerce").fillna(0.0)
    df["credit"] = pd.to_numeric(df["credit"], errors="coerce").fillna(0.0)
    df = df[df["trx_date"].notna()].copy()
    if df.empty:
        hints = extract_account_hints_from_raw(raw_df, bank, filename, sheet_name)
        status = pd.DataFrame(
            [create_status_row(bank, key, name, filename, sheet_name, "NO_TRANSACTION", "Tidak ada transaksi valid", 0) for key, name in hints]
        )
        return pd.DataFrame(), status, [f"{filename}[{sheet_name}]: tidak ada transaksi valid"]

    df["bank"] = bank
    df["source_file"] = filename
    df["source_sheet"] = sheet_name
    df["row_order"] = range(len(df))

    tx_cols = [
        "bank", "account_key", "account_name", "trx_date", "description", "amount", "dc",
        "debit", "credit", "balance", "opening_balance_explicit", "source_file", "source_sheet", "row_order",
    ]
    df = df[tx_cols]

    status_rows = []
    for account_key, group in df.groupby("account_key", dropna=False):
        names = [normalize_spaces(v) for v in group["account_name"].tolist() if normalize_spaces(v)]
        status_rows.append(
            create_status_row(
                bank,
                str(account_key),
                names[0] if names else "",
                filename,
                sheet_name,
                "OK",
                "Transaksi berhasil dibaca",
                len(group),
            )
        )

    notes.append(f"{filename}[{sheet_name}]: transaksi={len(df)}")
    return df, pd.DataFrame(status_rows), notes


def read_csv_any(file_bytes: bytes) -> pd.DataFrame:
    last_error = None
    for enc in ["utf-8", "utf-8-sig", "cp1252", "latin1"]:
        try:
            return pd.read_csv(io.BytesIO(file_bytes), header=None, encoding=enc)
        except Exception as exc:
            last_error = exc
    raise ValueError(f"Gagal membaca CSV: {last_error}")


def parse_tabular_file(file_bytes: bytes, filename: str, bank: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    ext = Path(filename).suffix.lower()
    notes: List[str] = []
    tx_list: List[pd.DataFrame] = []
    status_list: List[pd.DataFrame] = []

    if ext == ".csv":
        workbook = {"CSV": read_csv_any(file_bytes)}
    else:
        workbook = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)

    for sheet_name, raw_df in workbook.items():
        if bank.upper() == "MANDIRI":
            tx_df, status_df, current_notes = parse_mandiri_sheet(raw_df, filename, str(sheet_name), bank)
        else:
            tx_df, status_df, current_notes = parse_generic_tabular_sheet(raw_df, filename, str(sheet_name), bank)

        notes.extend(current_notes)
        if not tx_df.empty:
            tx_list.append(tx_df)
        if not status_df.empty:
            status_list.append(status_df)

    tx = pd.concat(tx_list, ignore_index=True) if tx_list else pd.DataFrame()
    status = pd.concat(status_list, ignore_index=True) if status_list else pd.DataFrame(columns=STATUS_COLUMNS)
    return tx, status, notes


def finalize_transactions(df: pd.DataFrame, bank: str, deduplicate: bool) -> pd.DataFrame:
    if df.empty:
        return df

    result = df.copy()
    result["account_key"] = result["account_key"].apply(lambda x: normalize_account_key(x, bank))
    result["account_name"] = result["account_name"].apply(normalize_spaces)
    result["description"] = result["description"].apply(normalize_spaces)
    result["dc"] = result["dc"].apply(standardize_dc)
    result["debit"] = pd.to_numeric(result["debit"], errors="coerce").fillna(0.0)
    result["credit"] = pd.to_numeric(result["credit"], errors="coerce").fillna(0.0)
    result["amount"] = pd.to_numeric(result["amount"], errors="coerce")
    result["balance"] = pd.to_numeric(result["balance"], errors="coerce")
    result["opening_balance_explicit"] = pd.to_numeric(result["opening_balance_explicit"], errors="coerce")
    result["trx_date"] = pd.to_datetime(result["trx_date"], errors="coerce").dt.normalize()

    if result["amount"].isna().any():
        result.loc[result["amount"].isna(), "amount"] = (
            result.loc[result["amount"].isna(), ["debit", "credit"]].max(axis=1)
        )

    result = result.sort_values(
        by=["account_key", "trx_date", "source_file", "source_sheet", "row_order"],
        kind="stable",
    ).reset_index(drop=True)

    if deduplicate:
        result = result.drop_duplicates(
            subset=["account_key", "trx_date", "description", "debit", "credit", "balance", "source_file"],
            keep="first",
        ).reset_index(drop=True)

    return result


def derive_opening_from_first_row(row: pd.Series) -> float:
    balance = row.get("balance")
    amount = row.get("amount")
    dc = standardize_dc(row.get("dc", ""))

    if pd.notna(balance) and pd.notna(amount) and amount is not None:
        if dc == "CR":
            return float(balance) - abs(float(amount))
        if dc == "DB":
            return float(balance) + abs(float(amount))

    debit = float(row.get("debit", 0) or 0)
    credit = float(row.get("credit", 0) or 0)
    if pd.notna(balance):
        if credit > 0:
            return float(balance) - credit
        if debit > 0:
            return float(balance) + debit
        return float(balance)

    explicit = row.get("opening_balance_explicit")
    if pd.notna(explicit):
        return float(explicit)

    return 0.0


def build_daily_summary(transactions: pd.DataFrame, status_df: pd.DataFrame, bank: str) -> pd.DataFrame:
    master_df = build_master_dataframe(bank)
    master_names = dict(zip(master_df["account_key"], master_df["account_name"])) if not master_df.empty else {}

    all_accounts = set(master_names.keys())
    if not status_df.empty:
        all_accounts.update(status_df["account_key"].astype(str))
    if not transactions.empty:
        all_accounts.update(transactions["account_key"].astype(str))

    if not all_accounts:
        return pd.DataFrame(
            columns=["Tanggal", "Rekening", "Nama Rekening", "Saldo Awal", "Debit", "Kredit", "Saldo Akhir", "Jumlah Transaksi"]
        )

    if transactions.empty:
        return pd.DataFrame(
            columns=["Tanggal", "Rekening", "Nama Rekening", "Saldo Awal", "Debit", "Kredit", "Saldo Akhir", "Jumlah Transaksi"]
        )

    min_date = transactions["trx_date"].min()
    max_date = transactions["trx_date"].max()
    date_range = pd.date_range(min_date, max_date, freq="D")

    tx = transactions.copy()
    tx["Tanggal"] = tx["trx_date"].dt.normalize()

    status_name_map = {}
    if not status_df.empty:
        grouped_names = status_df.groupby("account_key", dropna=False)["account_name"].apply(lambda s: next((normalize_spaces(v) for v in s if normalize_spaces(v)), ""))
        status_name_map = grouped_names.to_dict()

    records: List[Dict[str, object]] = []

    for account_key in sorted(all_accounts):
        account_rows = tx[tx["account_key"].astype(str) == str(account_key)].copy()
        account_rows = account_rows.sort_values(["Tanggal", "source_file", "source_sheet", "row_order"], kind="stable")

        name = master_names.get(account_key) or status_name_map.get(account_key, "")
        prev_close: Optional[float] = None

        explicit_opening = account_rows["opening_balance_explicit"].dropna()
        if not explicit_opening.empty:
            prev_close = float(explicit_opening.iloc[0])

        for current_date in date_range:
            day_rows = account_rows[account_rows["Tanggal"] == current_date].copy()
            if day_rows.empty:
                opening = prev_close if prev_close is not None else 0.0
                closing = opening
                debit = 0.0
                credit = 0.0
                count = 0
            else:
                day_rows = day_rows.sort_values(["row_order", "source_file", "source_sheet"], kind="stable")
                first_row = day_rows.iloc[0]
                last_row = day_rows.iloc[-1]

                opening = prev_close if prev_close is not None else derive_opening_from_first_row(first_row)
                debit = float(day_rows["debit"].sum())
                credit = float(day_rows["credit"].sum())

                if pd.notna(last_row["balance"]):
                    closing = float(last_row["balance"])
                else:
                    closing = opening + credit - debit

                count = len(day_rows)

            prev_close = closing
            records.append(
                {
                    "Tanggal": current_date.normalize(),
                    "Rekening": normalize_account_key(account_key, bank),
                    "Nama Rekening": name,
                    "Saldo Awal": float(opening),
                    "Debit": float(debit),
                    "Kredit": float(credit),
                    "Saldo Akhir": float(closing),
                    "Jumlah Transaksi": int(count),
                }
            )

    daily = pd.DataFrame(records)
    if daily.empty:
        return daily

    if not master_df.empty:
        order_map = dict(zip(master_df["account_key"], master_df["master_order"]))
        daily["sort_order"] = daily["Rekening"].map(order_map).fillna(999999)
    else:
        daily["sort_order"] = 999999

    daily = daily.sort_values(["Tanggal", "sort_order", "Rekening"], kind="stable").drop(columns=["sort_order"]).reset_index(drop=True)
    return daily


def build_rekap_from_daily(daily: pd.DataFrame, bank: str) -> pd.DataFrame:
    if daily.empty:
        columns = ["Rekening", "Nama Rekening", "Saldo Awal", "Debit", "Kredit", "Saldo Akhir", "Jumlah Transaksi"]
        return pd.DataFrame(columns=columns)

    grouped = daily.groupby("Rekening", dropna=False, sort=False)
    rows = []
    for account_key, group in grouped:
        group = group.sort_values("Tanggal", kind="stable")
        rows.append(
            {
                "Rekening": account_key,
                "Nama Rekening": next((normalize_spaces(v) for v in group["Nama Rekening"] if normalize_spaces(v)), ""),
                "Saldo Awal": float(group.iloc[0]["Saldo Awal"]),
                "Debit": float(group["Debit"].sum()),
                "Kredit": float(group["Kredit"].sum()),
                "Saldo Akhir": float(group.iloc[-1]["Saldo Akhir"]),
                "Jumlah Transaksi": int(group["Jumlah Transaksi"].sum()),
            }
        )

    rekap = pd.DataFrame(rows)
    master_df = build_master_dataframe(bank)
    if not master_df.empty:
        rekap = rekap.merge(master_df, how="left", left_on="Rekening", right_on="account_key")
        rekap["Nama Rekening"] = rekap["account_name"].combine_first(rekap["Nama Rekening"])
        rekap["master_order"] = rekap["master_order"].fillna(999999)
        rekap = rekap.drop(columns=["account_key", "account_name"]).sort_values(["master_order", "Rekening"], kind="stable")
        rekap = rekap.drop(columns=["master_order"])
    else:
        rekap = rekap.sort_values("Rekening", kind="stable")

    return rekap.reset_index(drop=True)


def make_display_copy(df: pd.DataFrame, money_cols: List[str]) -> pd.DataFrame:
    result = df.copy()
    for col in money_cols:
        if col in result.columns:
            result[col] = result[col].apply(format_currency)
    return result


def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    clean = re.sub(r"[:\\/?*\[\]]", "_", str(name)).strip()
    clean = clean[:31] or "Sheet"
    if clean not in used_names:
        used_names.add(clean)
        return clean

    base = clean[:28] or "Sht"
    counter = 1
    while True:
        candidate = f"{base}_{counter}"[:31]
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def autosize_worksheet(ws) -> None:
    for col_cells in ws.columns:
        length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[letter].width = min(max(length + 2, 12), 40)


def build_excel_export(rekap_df: pd.DataFrame, daily_df: pd.DataFrame, detail_df: pd.DataFrame, status_df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    used_names: set[str] = set()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        rekap_df.to_excel(writer, sheet_name=sanitize_sheet_name("Rekap", used_names), index=False)

        detail_export = detail_df.copy()
        if "Tanggal" in detail_export.columns:
            detail_export["Tanggal"] = pd.to_datetime(detail_export["Tanggal"], errors="coerce").dt.strftime("%Y-%m-%d")
        detail_export.to_excel(writer, sheet_name=sanitize_sheet_name("Semua_Transaksi", used_names), index=False)

        status_export = status_df.rename(
            columns={
                "account_key": "Rekening",
                "account_name": "Nama Rekening",
                "source_file": "File",
                "source_sheet": "Sheet",
                "parse_status": "Status",
                "parse_note": "Catatan",
                "transaction_count": "Jumlah Transaksi",
            }
        )
        status_export.to_excel(writer, sheet_name=sanitize_sheet_name("Status_File", used_names), index=False)

        if not daily_df.empty:
            for current_date, group in daily_df.groupby("Tanggal", sort=True):
                export_group = group[["Rekening", "Saldo Awal", "Debit", "Kredit", "Saldo Akhir"]].copy()
                export_group.to_excel(
                    writer,
                    sheet_name=sanitize_sheet_name(pd.Timestamp(current_date).strftime("%Y-%m-%d"), used_names),
                    index=False,
                )

        workbook = writer.book
        money_cols = {"Saldo Awal", "Debit", "Kredit", "Saldo Akhir", "Saldo"}
        for ws in workbook.worksheets:
            headers = [cell.value for cell in ws[1]]
            for idx, header in enumerate(headers, start=1):
                if header in money_cols:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=idx).number_format = "#,##0.00"
            autosize_worksheet(ws)

    output.seek(0)
    return output.getvalue()


def parse_uploaded_files(uploaded_files: List, bank: str, deduplicate: bool) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], List[str]]:
    tx_list: List[pd.DataFrame] = []
    status_list: List[pd.DataFrame] = []
    notes: List[str] = []
    errors: List[str] = []

    total = len(uploaded_files)
    progress = st.progress(0, text="Memproses file...")

    for idx, uploaded in enumerate(uploaded_files, start=1):
        filename = uploaded.name
        ext = Path(filename).suffix.lower()
        payload = uploaded.getvalue()

        try:
            if ext == ".pdf" and bank.upper() == "BCA":
                tx_df, status_df, current_notes = parse_bca_pdf(payload, filename, bank)
            elif ext in {".csv", ".xlsx", ".xls"}:
                tx_df, status_df, current_notes = parse_tabular_file(payload, filename, bank)
            else:
                tx_df = pd.DataFrame()
                status_df = pd.DataFrame([create_status_row(bank, "", "", filename, "FILE", "UNSUPPORTED", "Format file tidak didukung", 0)])
                current_notes = [f"{filename}: format tidak didukung"]

            notes.extend(current_notes)
            if not tx_df.empty:
                tx_list.append(tx_df)
            if not status_df.empty:
                status_list.append(status_df)
            if tx_df.empty:
                errors.append(f"{filename}: tidak ada transaksi valid, tetapi file tetap dicatat")
        except Exception as exc:
            errors.append(f"{filename}: error - {exc}")
            status_list.append(pd.DataFrame([create_status_row(bank, "", "", filename, "FILE", "ERROR", str(exc), 0)]))

        progress.progress(idx / total, text=f"Memproses file {idx}/{total}: {filename}")

    progress.empty()

    transactions = pd.concat(tx_list, ignore_index=True) if tx_list else pd.DataFrame()
    status_df = pd.concat(status_list, ignore_index=True) if status_list else pd.DataFrame(columns=STATUS_COLUMNS)

    if not transactions.empty:
        transactions = finalize_transactions(transactions, bank, deduplicate)

    return transactions, status_df, notes, errors


def build_detail_transactions(transactions: pd.DataFrame) -> pd.DataFrame:
    if transactions.empty:
        return pd.DataFrame(columns=["Rekening", "Nama Rekening", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo", "File", "Sheet"])

    detail = transactions.rename(
        columns={
            "account_key": "Rekening",
            "account_name": "Nama Rekening",
            "trx_date": "Tanggal",
            "description": "Keterangan",
            "debit": "Debit",
            "credit": "Kredit",
            "balance": "Saldo",
            "source_file": "File",
            "source_sheet": "Sheet",
        }
    )[["Rekening", "Nama Rekening", "Tanggal", "Keterangan", "Debit", "Kredit", "Saldo", "File", "Sheet"]].copy()

    return detail.sort_values(["Rekening", "Tanggal", "File", "Sheet"], kind="stable").reset_index(drop=True)


def main() -> None:
    st.title("Pembaca Rekening Koran")
    st.caption("BCA, Mandiri, BNI, atau bank lainnya.")

    with st.sidebar:
        bank = st.selectbox("Bank", ["BCA", "MANDIRI", "BNI", "LAINNYA"], index=0)
        deduplicate = st.checkbox("Hapus duplikat transaksi identik", value=True)

        if bank == "BCA":
            st.info("Master rekening BCA aktif sebagai default di dalam code.")
        elif bank == "MANDIRI":
            st.info("Format Mandiri membaca kolom B/E/J/L/P dari rekening koran yang Anda kirim.")
        else:
            st.info("Untuk bank selain BCA/Mandiri, gunakan file Excel/CSV dengan header standar.")

    uploaded_files = st.file_uploader(
        "Upload rekening koran / mutasi",
        type=["pdf", "csv", "xlsx", "xls"],
        accept_multiple_files=True,
    )

    if not uploaded_files:
        st.info("Upload file dulu.")
        return

    transactions, status_df, notes, errors = parse_uploaded_files(uploaded_files, bank, deduplicate)

    if status_df.empty:
        st.error("Tidak ada file yang berhasil dicatat.")
        return

    daily_df = build_daily_summary(transactions, status_df, bank)
    rekap_df = build_rekap_from_daily(daily_df, bank)
    detail_df = build_detail_transactions(transactions)

    st.success(
        f"Upload: {len(uploaded_files)} file | "
        f"Tercatat: {status_df['source_file'].nunique()} file | "
        f"Dengan transaksi: {transactions['source_file'].nunique() if not transactions.empty else 0} file"
    )

    st.subheader("Rekap per Rekening")
    st.dataframe(
        make_display_copy(rekap_df, ["Saldo Awal", "Debit", "Kredit", "Saldo Akhir"]),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Status File / Sheet")
    status_view = status_df.rename(
        columns={
            "account_key": "Rekening",
            "account_name": "Nama Rekening",
            "source_file": "File",
            "source_sheet": "Sheet",
            "parse_status": "Status",
            "parse_note": "Catatan",
            "transaction_count": "Jumlah Transaksi",
        }
    )
    st.dataframe(status_view, use_container_width=True, hide_index=True)

    st.subheader("Detail Transaksi")
    st.dataframe(
        make_display_copy(detail_df, ["Debit", "Kredit", "Saldo"]),
        use_container_width=True,
        hide_index=True,
    )

    excel_bytes = build_excel_export(rekap_df, daily_df, detail_df, status_df)
    st.download_button(
        "Download Excel",
        data=excel_bytes,
        file_name=f"rekap_{bank.lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.download_button(
        "Download Rekap CSV",
        data=rekap_df.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"rekap_{bank.lower()}.csv",
        mime="text/csv",
    )

    st.download_button(
        "Download Detail CSV",
        data=detail_df.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"detail_{bank.lower()}.csv",
        mime="text/csv",
    )

    with st.expander("Log parser"):
        for note in notes:
            st.write(f"- {note}")
        for err in errors:
            st.write(f"- {err}")


if __name__ == "__main__":
    main()
